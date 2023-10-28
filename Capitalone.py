authJWT = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiJ9.eyJuYmYiOjE2OTYwMzIwMDAsImFwaV9zdWIiOiJhMmRlMjUxNGRmZTAwYjM5OTFhNzg0NTY4NzJkZDJjOTE1YWY5NjgzOTc2ZDAyZTdiZjk5NTIzMWE2MDllMDVjMTcxNzIwMDAwMDAwMCIsInBsYyI6IjVkY2VjNzRhZTk3NzAxMGUwM2FkNjQ5NSIsImV4cCI6MTcxNzIwMDAwMCwiZGV2ZWxvcGVyX2lkIjoiYTJkZTI1MTRkZmUwMGIzOTkxYTc4NDU2ODcyZGQyYzkxNWFmOTY4Mzk3NmQwMmU3YmY5OTUyMzFhNjA5ZTA1YyJ9.NiTvZUyTERVaV7krUmkiRgXbXhAeWEEZHkaus8aJ5pTMG4RYbZSP5reayEl1xi8kHO9g8jOWkj6v8ygUqyuPGcsL2jtqTMen1kUhe5m2bO6qQmdOvF6leyQkf7qWeYC8pzcIYhCnII5ZAcJxs3Lx3F3o6Zqy3aLsP1OxmG_TWiyWfdfhdCO8dekvQxXnp61oYAT8BvMPTbvoh6g9xPPO3q5KyYEAeLDGiPNeFML6CaDpvrwH_gf_ilDZJ84jGXuRGN5q_BqUgJq8QXahrbhvzM1ewNsC48Ljt0jFVD1kErMn4AVzJUkpZhzT6RJUvWC0e-lVObJ85yDCcbNdKOn7Dw'
final_balance = list()
full_dictionarry = {'Transactions':[]}

import json
import requests
from collections import Counter





###Create a random Account

##authJWT='eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiJ9.eyJuYmYiOjE2OTYwMzIwMDAsImFwaV9zdWIiOiI4MjYyMThkNmIzZDU3MTkxM2FiZmExZjY3ZDdmYTRjOWQ5NDM1YjNmODAyZDAxYTFjOTA1NTk1MTc0YWQ2N2Q0MTcxNzIwMDAwMDAwMCIsInBsYyI6IjVkY2VjNzRhZTk3NzAxMGUwM2FkNjQ5NSIsImV4cCI6MTcxNzIwMDAwMCwiZGV2ZWxvcGVyX2lkIjoiODI2MjE4ZDZiM2Q1NzE5MTNhYmZhMWY2N2Q3ZmE0YzlkOTQzNWIzZjgwMmQwMWExYzkwNTU5NTE3NGFkNjdkNCJ9.YmNaE4xkJ9sK2yoVSkQzhALV4NNrkZx0MKL55m3Wn2az05DznVxAxYYqioNXNdLySIJlpZRxbbhPGd1Mu9xB6pGlmGuffzECQS-Wjk4orpx87pNh-tdQTIL-1kjgvRjv00Gr7yzWgLea1j3b4xw4BYq86f7ikXkPqLoE_iBjsU75Tl03UrfibDjH9ChgOBMVHMH33pAz2pFr7hIHvhg6xRR1O_s08mZsr51yb0bucase2hUqS0Whh1VfuokvaAbvDrPYLLEEBMRnYIqE5s15QjSrHmywofD_Yy0nr0Uy_-ihuXOrOQ49G80Gs_Enw_sCQUFwbgEH7i1oivn5ft7_pQ'

##import json
##import requests
# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *



import tkinter as tk
from tkinter import ttk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
 


# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\N1207727\\Projects\\HackDuck\\excel.xlsx')

# create the sheet object
sheet = wb.active



def generate_the_data():
    global full_dictionarry
    headers = {
    'Authorization': f'Bearer {authJWT}',
    'Content-Type': 'application/json',
    'version': '1.0'
    }
    for i in range(1):
        quantity = 25
        payload = json.dumps({"quantity": quantity})
        account_id = "95526223"

        response = requests.post(f"https://sandbox.capitalone.co.uk/developer-services-platform-pr/api/data/transactions/accounts/{account_id}/create", headers=headers, data=payload).text
        json_response = json.loads(response)

        final_balance.append(json_response["newBalance"])
        json_response.pop('newBalance')
        for x in range(len(json_response["Transactions"])):
            if json_response["Transactions"][x]['status'] != "Successful":
                continue
            else:
                # if the transaction is fine and accepted then we will add it to the dictionary to build the database we want and test on it ... we append the (transaction [x] because it is a list of a transactions so its indix is x ) ... 
                full_dictionarry["Transactions"].append(json_response["Transactions"][x])
def categories_list():
    categories_list = []
    
    # Loop through the list of dictionaries and extract the 'category' key
    for i in full_dictionarry['Transactions']:
        # Check if 'merchant' key exists in the transaction dictionary
        if 'merchant' in i and 'category' in i['merchant']:
            # Append the 'category' value to the categories_list
            categories_list.append(i['merchant']['category'])
    
    sorted_dict = Counter(categories_list)
    return sorted_dict
def ploting_categories(sorted_dict):
    import matplotlib.pyplot as plt

    # Extract category names and their counts
    categories = list(sorted_dict.keys())
    counts = list(sorted_dict.values())

    # Create a bar chart
    plt.figure(figsize=(10, 6))
    plt.barh(categories, counts, color='skyblue')
    plt.xlabel('Number of Occurrences')
    plt.ylabel('Categories')
    plt.title('Occurrences of Categories')
    plt.gca().invert_yaxis()  # Invert the y-axis for better readability
    
    
    plt.tight_layout()

    # Save the plot as an image (optional)
    plt.savefig('categories_bar_chart.png')

    # Show the plot
    plt.show()


    pass
generate_the_data()



def user_question():
    food_and_Dining_budget = float(input('Enter how much you want to spend on Food and Dining'))
    bills_Utilities_budget  = float(input('Enter how much you want to spend on Bills and Utilities'))
    auto_Transport_budget = float(input('Enter how much you want to spend on Auto and Transport'))
    shopping_budget = float(input('Enter how much you want to spend on Shopping'))
    other_budget = float(input('Enter how much you want to spend on other'))
def sorted_amount():
    sorted_amount_list = [(transaction['amount'], transaction['creditDebitIndicator'], transaction['merchant']['category'])
                  for transaction in full_dictionarry['Transactions']]
    return sorted_amount_list
def categorize_transactions(sorted_amount_list):
    categories_data = {}  # Dictionary to store category-wise spending information

    for x in sorted_amount_list:
        amount, credit_debit, category = x[0], x[1], x[2] 
        (29.36, 'Debit', 'Education')
        # Check if the category is already in the dictionary, if not, create it
        if category not in categories_data:
            categories_data[category] = {
                'total_amount': 0,
                'debit_amount': 0,
                'credit_amount': 0
            }

        # Update total amount for the category
        categories_data[category]['total_amount'] += amount

        # Categorize transactions into debit and credit
        if credit_debit == 'Debit':
            categories_data[category]['debit_amount'] += amount
        elif credit_debit == 'Credit':
            categories_data[category]['credit_amount'] += amount

    return categories_data

print(categorize_transactions(sorted_amount_list=sorted_amount()))


#def GraphPlot():
#	# Create the main application window
#	root = tk.Tk()
#	root.title("Bank Transaction Analyzer")
 
#	# Load your data and perform analysis here
 
#	# Create a Matplotlib figure
#	fig = Figure(figsize=(6, 4), dpi=100)
#	ax = fig.add_subplot(111)
 
#	# Create a bar chart using Matplotlib
#	data = [10, 20, 30, 40, 50]  # Replace with your actual data
#	labels = ['Category A', 'Category B', 'Category C', 'Category D', 'Category E']
#	ax.bar(labels, data)
 
#	# Embed the Matplotlib figure in the Tkinter window
#	canvas = FigureCanvasTkAgg(fig, master=root)
#	canvas_widget = canvas.get_tk_widget()
#	canvas_widget.pack()
 
#	root.mainloop()
#GraphPlot()
 




def excel():
	
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Enter how much do you want to spend on Entertainment?"
	sheet.cell(row=1, column=2).value = "Enter how much do you want to spend on Education?"
	sheet.cell(row=1, column=3).value = "Enter how much do you want to spend on Shopping?"
	sheet.cell(row=1, column=4).value = "Enter how much do you want to spend on Personal Care?"
	sheet.cell(row=1, column=5).value = "Enter how much do you want to spend on Health and Fitness?"
	sheet.cell(row=1, column=6).value = "Enter how much do you want to spend on Food and Dining?"
	sheet.cell(row=1, column=7).value = "Enter how much do you want to spend on Gifts and Donations?"


# Function to set focus (cursor)
def focus1(event):
	# set focus on the course_field box
	course_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	sem_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	form_no_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	contact_no_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	email_id_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the address_field box
	address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)


# Function to take data from GUI 
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=7).value = address_field.get()

		# save the file
		wb.save('C:\\Users\\N1207727\\Projects\\HackDuck\\excel.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()


# Driver code
if __name__ == "__main__":
	
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='White')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("700x500")

	excel()

	# create a Form label
	heading = Label(root, text="Capital Applictaion Details Form", bg="White")

	# create a Name label
	name = Label(root, text="Enter how much do you want to spend on Entertainment?", bg="White")

	# create a Course label
	course = Label(root, text="Enter how much do you want to spend on Education?", bg="White")

	# create a Semester label
	sem = Label(root, text="Enter how much do you want to spend on Shopping?", bg="White")

	# create a Form No. label
	form_no = Label(root, text="Enter how much do you want to spend on Personal Care?", bg="White")

	# create a Contact No. label
	contact_no = Label(root, text="Enter how much do you want to spend on Health and Fitness?", bg="White")

	# create a Email id label
	email_id = Label(root, text="Enter how much do you want to spend on Food and Dining?", bg="White")

	# create a address label
	address = Label(root, text="Enter how much do you want to spend on Gifts and Donations?", bg="White")

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	heading.grid(row=0, column=1)
	name.grid(row=1, column=0)
	course.grid(row=2, column=0)
	sem.grid(row=3, column=0)
	form_no.grid(row=4, column=0)
	contact_no.grid(row=5, column=0)
	email_id.grid(row=6, column=0)
	address.grid(row=7, column=0)

	# create a text entry box
	# for typing the information
	name_field = Entry(root)
	course_field = Entry(root)
	sem_field = Entry(root)
	form_no_field = Entry(root)
	contact_no_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)

	# bind method of widget is used for
	# the binding the function with the events

	# whenever the enter key is pressed
	# then call the focus1 function
	name_field.bind("<Return>", focus1)

	# whenever the enter key is pressed
	# then call the focus2 function
	course_field.bind("<Return>", focus2)

	# whenever the enter key is pressed
	# then call the focus3 function
	sem_field.bind("<Return>", focus3)

	# whenever the enter key is pressed
	# then call the focus4 function
	form_no_field.bind("<Return>", focus4)

	# whenever the enter key is pressed
	# then call the focus5 function
	contact_no_field.bind("<Return>", focus5)

	# whenever the enter key is pressed
	# then call the focus6 function
	email_id_field.bind("<Return>", focus6)

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	name_field.grid(row=1, column=1, ipadx="100")
	course_field.grid(row=2, column=1, ipadx="100")
	sem_field.grid(row=3, column=1, ipadx="100")
	form_no_field.grid(row=4, column=1, ipadx="100")
	contact_no_field.grid(row=5, column=1, ipadx="100")
	email_id_field.grid(row=6, column=1, ipadx="100")
	address_field.grid(row=7, column=1, ipadx="100")

	# call excel function
	excel()

	# create a Submit Button and place into the root window
	submit = Button(root, text="Submit", fg="Black",
							bg="White", command=ploting_categories)

	submit.grid(row=8, column=1)


	# start the GUI
	root.mainloop()




